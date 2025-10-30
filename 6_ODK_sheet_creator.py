import pandas as pd
import os
import openpyxl
import re
import datetime

# ================================================================
# STEP 1 - Create choices sheet from ODK Only lists and concepts
# ================================================================

# Input and output folders
input_path = r"D:\University of Cambridge\ARCH_MAHSA - General\MAHSA_Database\Thesauri\Thesauri_Audit\Spreadsheets\1_Processing\excel_thesauri_ODK_only.xlsx"
output_folder = r"D:\University of Cambridge\ARCH_MAHSA - General\MAHSA_Database\Thesauri\Thesauri_Audit\Spreadsheets\5_Updated_ODK_form\Choices_sheets"

# --- Create dated subfolder for today's outputs ---
today_str = datetime.date.today().strftime("%Y%m%d")  # e.g. 20251016
dated_output_folder = os.path.join(output_folder, today_str)

# Create the folder if it doesn't already exist
os.makedirs(dated_output_folder, exist_ok=True)

# Redirect all output paths to this folder
output_folder = dated_output_folder
print(f"📁 Output folder set to: {output_folder}")

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Prepare new sheet
output_data = [["list_name", "name", "label", "media::image", "transect_method_list", "Institute name", "heritage_resource_classification"]]

current_list_name = None
header_row = None
header_map = {}

for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    cleaned_row = [str(c).strip() if c is not None else "" for c in row]
    first_cell = cleaned_row[0].lower() if cleaned_row else ""

    if first_cell.startswith("odk list name"):
        current_list_name = cleaned_row[1] if len(cleaned_row) > 1 else ""
        header_row = None
        header_map = {}
        continue

    if first_cell == "odk name" or first_cell == "odk value":
        header_row = cleaned_row
        for idx, col_name in enumerate(header_row):
            header_map[col_name.lower()] = idx
        continue

    if not current_list_name or not header_row or not any(cleaned_row):
        continue

    def get_val(col_key):
        idx = header_map.get(col_key.lower())
        return cleaned_row[idx] if idx is not None and idx < len(cleaned_row) else ""

    name_val = get_val("odk name") or get_val("odk value")
    label_val = get_val("odk label") or get_val("odk term")
    image_val = get_val("odk image file")
    filter_val = get_val("odk filter")
    multi_val = get_val("odk multi list")

    if not name_val and not label_val:
        continue

    # Split comma-separated multi values into separate rows
    multi_vals = [v.strip() for v in multi_val.split(",") if v.strip()] if multi_val else [""]

    for mv in multi_vals:
        output_data.append([
            current_list_name,
            name_val,
            label_val,
            image_val,
            filter_val,
            "",  # Institute name blank
            mv   # each multi_val entry
        ])

# Save Step 1 output
odk_only_path = os.path.join(output_folder, "ODK_only_concepts.xlsx")
os.makedirs(output_folder, exist_ok=True)
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "ODK Concepts"
for row in output_data:
    ws_out.append(row)
wb_out.save(odk_only_path)
print(f"✅ ODK Only concepts saved as: {odk_only_path}")

# ================================================================
# STEP 2 - Create choices sheet from PO details (Bulk Import)
# ================================================================

bulk_input = r"D:\University of Cambridge\ARCH_MAHSA - General\MAHSA_Database\ArchesDataDigitization\CommonDataSheets\Common_BulkImportSheet.xlsx"
bulk_output = os.path.join(output_folder, "ODK_PO_entries.xlsx")

df_raw = pd.read_excel(bulk_input, sheet_name="Person-Organization RM", skiprows=3)
df_filtered = df_raw[df_raw['KOBO Account'].astype(str).str.strip().str.lower() == 'yes']

df = pd.DataFrame({
    'name': df_filtered['MAHSA_ID'],
    'label': df_filtered['Name'],
    'po_institution': df_filtered['Related Organization']
})

# Map institute names
po_to_name = pd.Series(df_raw['Name'].values, index=df_raw['MAHSA_ID']).to_dict()
po_to_odk = pd.Series(df_raw['ODK Institute Name'].values, index=df_raw['MAHSA_ID']).to_dict()
df['institutename'] = df['po_institution'].map(po_to_name)
df['institute_name'] = df['po_institution'].map(po_to_odk)

df.sort_values(by=['institutename', 'label'], inplace=True, ignore_index=True)

# Add "Not listed" row per institute
new_rows = []
for name, group in df.groupby('institutename', sort=False):
    new_rows.append(group)
    not_listed = {
        'name': 'not_listed',
        'label': 'Not listed here',
        'po_institution': None,
        'institutename': name,
        'institute_name': group['institute_name'].iloc[0] if not group['institute_name'].isna().all() else None
    }
    new_rows.append(pd.DataFrame([not_listed]))

df = pd.concat(new_rows, ignore_index=True)
df.insert(0, 'list_name', 'recorder_list')

df_unique = df[['institutename', 'institute_name']].drop_duplicates().rename(
    columns={'institutename': 'label', 'institute_name': 'name'}
)
df_unique.insert(0, 'list_name', 'institute_name')

df = df[['list_name', 'name', 'label', 'institute_name']]
df_combined = pd.concat([df, df_unique], ignore_index=True, sort=False)
df_combined['media::image'] = ""
df_combined['transect_method_list'] = ""
df_combined['heritage_resource_classification'] = ""

df_combined = df_combined[['list_name', 'name', 'label', 'media::image', 'transect_method_list',
                           'institute_name', 'heritage_resource_classification']]
df_combined.to_excel(bulk_output, index=False)
print(f"✅ Bulk Import choices saved as: {bulk_output}")

# ================================================================
# STEP 3 - Create choices from Complete Thesauri Concepts
# ================================================================

complete_concepts_dir = r"D:\University of Cambridge\ARCH_MAHSA - General\MAHSA_Database\Thesauri\Thesauri_Audit\Spreadsheets\3_Complete_concepts"
csv_pattern = re.compile(r"complete_thesauri_concepts_(\d{8})\.csv$")
csv_candidates = [(f, m.group(1)) for f in os.listdir(complete_concepts_dir) if (m := csv_pattern.match(f))]
if not csv_candidates:
    raise FileNotFoundError("No complete_thesauri_concepts_*.csv found.")
csv_candidates.sort(key=lambda x: x[1])
csv_name, _ = csv_candidates[-1]
csv_path = os.path.join(complete_concepts_dir, csv_name)
print(f"📘 Using most recent thesauri file: {csv_name}")

# Load and keep only relevant columns
df_thes = pd.read_csv(csv_path, usecols=["ODK_list_name", "odk_value", "concept_key", "concept_value", "ODK_multi", "list_order"])

# --- Keep only rows with a valid, non-empty odk_value ---
df_thes = df_thes[
    df_thes["odk_value"].notna() &  # not NaN
    (df_thes["odk_value"].astype(str).str.strip().ne("")) &  # not empty string
    (df_thes["odk_value"].astype(str).str.lower().ne("nan"))  # not literal "nan"
]


# Expand comma-separated ODK_list_name into multiple rows

df_expanded_list = []
for _, row in df_thes.iterrows():
    list_names = [v.strip() for v in str(row["ODK_list_name"]).split(",") if v.strip()] if pd.notna(row["ODK_list_name"]) else [""]
    for ln in list_names:
        new_row = row.copy()
        new_row["ODK_list_name"] = ln
        df_expanded_list.append(new_row)

df_thes = pd.DataFrame(df_expanded_list)

# Expand comma-separated ODK_multi into multiple rows
df_expanded = []
for _, row in df_thes.iterrows():
    multi_vals = [v.strip() for v in str(row["ODK_multi"]).split(",") if v.strip()] if pd.notna(row["ODK_multi"]) else [""]
    for mv in multi_vals:
        new_row = row.copy()
        new_row["multi_val"] = mv
        df_expanded.append(new_row)
df_thes = pd.DataFrame(df_expanded)

# Sort by ODK_list_name, ODK_multi, concept_key, then list_order (if present)
df_thes = df_thes.sort_values(by=["ODK_list_name", "multi_val", "list_order", "concept_value"], na_position="last")

# Map to final structure
df_thes_final = pd.DataFrame({
    "list_name": df_thes["ODK_list_name"],
    "name": df_thes["odk_value"],
    "label": df_thes["concept_key"],
    "media::image": "",
    "transect_method_list": "",
    "institute_name": "",
    "heritage_resource_classification": df_thes["multi_val"]
})

# Save to Excel
thesauri_output = os.path.join(output_folder, "ODK_thesauri_concepts.xlsx")
df_thes_final.to_excel(thesauri_output, index=False)
print(f"✅ Thesauri concepts saved as: {thesauri_output}")

# ================================================================
# STEP 4 - Combine all three outputs (Thesauri, ODK Only, PO Entries)
# ================================================================

# Load all three
df_thes = pd.read_excel(thesauri_output)
df_odk = pd.read_excel(odk_only_path)
df_po = pd.read_excel(bulk_output)

# Normalize column names
def norm(df):
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    return df

df_thes, df_odk, df_po = map(norm, [df_thes, df_odk, df_po])

final_cols = ["list_name", "name", "label", "media::image", "transect_method_list", "institute_name", "heritage_resource_classification"]
df_thes = df_thes[final_cols]
df_odk = df_odk[final_cols]
df_po = df_po[final_cols]

# --- Merge logic ---
# 1. Start with thesauri (sorted already)
combined = df_thes.copy()

# 2. Append ODK Only concepts, preserving multi_val grouping
for _, row in df_odk.iterrows():
    list_name = row["list_name"]
    multi_val = row["heritage_resource_classification"]

    mask_list = combined["list_name"] == list_name
    mask_multi = combined["heritage_resource_classification"] == multi_val

    if mask_list.any():
        # If matching multi_val exists, place after the last occurrence
        if mask_multi.any():
            insert_idx = combined[mask_multi].index[-1] + 1
        else:
            insert_idx = combined[mask_list].index[-1] + 1
        combined = pd.concat(
            [combined.iloc[:insert_idx], pd.DataFrame([row]), combined.iloc[insert_idx:]],
            ignore_index=True
        )
    else:
        # If no matching list_name, append to end
        combined = pd.concat([combined, pd.DataFrame([row])], ignore_index=True)

# 3. Append PO entries last
combined = pd.concat([combined, df_po], ignore_index=True)

# Save final combined output
final_output = os.path.join(output_folder, "ODK_combined_concepts.xlsx")
combined.to_excel(final_output, index=False)
print(f"✅ All three datasets combined successfully!\n💾 Saved as: {final_output}")

# ================================================================
# STEP 5 - Update the Master ODK site form with new choices sheet
# ================================================================
import re
from openpyxl import load_workbook
import datetime

master_folder = r"D:\University of Cambridge\ARCH_MAHSA - General\MAHSA_Database\Thesauri\Thesauri_Audit\Spreadsheets\5_Updated_ODK_form\Master_ODK_site_form"
pattern = re.compile(r"MAHSA_Site_Form_V21_(\d{8})_(\d+)\.xlsx$", re.IGNORECASE)

# --- Find all master files ---
candidates = []
for f in os.listdir(master_folder):
    m = pattern.match(f)
    if m:
        candidates.append((f, int(m.group(2))))  # keep filename and N

if not candidates:
    raise FileNotFoundError("No master form found in Master_ODK_site_form folder.")

# Sort by N to find the highest
candidates.sort(key=lambda x: x[1])
latest_file, latest_num = candidates[-1]
latest_path = os.path.join(master_folder, latest_file)
print(f"📄 Latest master form found: {latest_file}")

# --- Always increment N ---
new_num = latest_num + 1
today_str = datetime.date.today().strftime("%Y%m%d")
new_filename = f"MAHSA_Site_Form_V21_{today_str}_{new_num}.xlsx"
new_path = os.path.join(master_folder, new_filename)

# --- Load the workbook and choices data ---
wb = load_workbook(latest_path)
if "choices" not in wb.sheetnames:
    raise ValueError("The workbook does not contain a sheet named 'choices'.")

ws = wb["choices"]

# --- Clear the choices sheet ---
ws.delete_rows(1, ws.max_row)

# --- Load combined ODK data ---
df_combined = combined

# --- Write headers first ---
for c_idx, col_name in enumerate(df_combined.columns, start=1):
    ws.cell(row=1, column=c_idx, value=col_name)

# --- Write data rows starting from row 2 ---
for r_idx, row in enumerate(df_combined.itertuples(index=False), start=2):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# --- Save new workbook ---
wb.save(new_path)
print(f"✅ Updated master form saved as: {new_filename}")
print(f"📂 Location: {new_path}")