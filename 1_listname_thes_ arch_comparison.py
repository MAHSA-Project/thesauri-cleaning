# Import necessary libraries
import pandas as pd
import os
import csv
import openpyxl
import numpy as np
from difflib import get_close_matches

# Define the data directory
data_dir = os.path.join(os.getcwd(), "D:/University of Cambridge/ARCH_MAHSA - General/MAHSA_Database/Thesauri/Thesauri_Audit/Spreadsheets/")

# Load the Excel workbook using openpyxl
workbook_path = os.path.join('D:/University of Cambridge/ARCH_MAHSA - General/MAHSA_Database/Thesauri/MAHSA_Thesauri_v5.xlsx')

workbook = openpyxl.load_workbook(workbook_path)

# Print all sheet names for reference
print("Original sheets:", workbook.sheetnames)

# Save the "ODK Only" sheet separately (to be used when generating new ODK form)
if 'ODK Only' in workbook.sheetnames:
    odk_sheet = workbook['ODK Only']

    # Create a new workbook and copy the ODK Only sheet into it
    odk_wb = openpyxl.Workbook()
    odk_ws = odk_wb.active
    odk_ws.title = 'ODK Only'

    # Copy contents cell by cell
    for row in odk_sheet.iter_rows(values_only=False):
        for cell in row:
            odk_ws[cell.coordinate].value = cell.value

    # Save this workbook separately
    odk_only_path = os.path.join(data_dir, '1_Processing/excel_thesauri_ODK_only.xlsx')
    odk_wb.save(odk_only_path)
    print(f"Saved 'ODK Only' sheet to: {odk_only_path}")

# Delete unnecessary sheets
sheets_to_delete = [
    'Temp Concept Sheet', 'Relationships', 'ODK Only', 'Guidelines',
    'TempWorkSheet', 'PalaeolithicChronology (in prg)'
]
for sheet in sheets_to_delete:
    del workbook[sheet]

# Save the cleaned workbook
processed_path = os.path.join(data_dir, '1_Processing/excel_thesauri_processed.xlsx')
workbook.save(processed_path)

# Read the processed workbook
xls = pd.ExcelFile(processed_path)
print("Sheets to process:", xls.sheet_names)

# Collect all sheets into a list of DataFrames
df_list = [pd.read_excel(processed_path, sheet_name=sheet, header=None) for sheet in xls.sheet_names]

# Concatenate all sheets into a single DataFrame and remove any columns after column 7
df = pd.concat(df_list, ignore_index=True)
df = df.iloc[:, :8]

# Create column 8: if column 0 is 'Resource Model Node', copy column 1; else empty string
df[8] = np.where(df[0] == 'Resource Model Node', df[1], '')

# Forward-fill missing values in column 8
df[8] = df[8].replace('', pd.NA).ffill()

# Copy column 1 to column 9
df[9] = df[1]

# Create col10 with BI Name from thesauri where present, else NA
df[10] = np.where(df[0] == 'BI Name', df[1], pd.NA)
df[11] = np.where(df[0] == 'ODK List Name', df[1], pd.NA)

# Forward-fill col10 within each group of col8
df[10] = (
    df.groupby(df[8])[10]
      .apply(lambda g: g.ffill())
      .reset_index(level=0, drop=True)
)

# Forward-fill col11 within each group of col8
df[11] = (
    df.groupby(df[8])[11]
      .apply(lambda g: g.ffill())
      .reset_index(level=0, drop=True)
)

# Fallback — if col10 is still empty/NaN, copy col8
df[10] = df[10].fillna(df[8])

# Replace blank strings in column 0 with NaN
df[0] = df[0].replace('', pd.NA)

# Drop rows where column 0 is in specific labels or both column 0 and 1 are NaN
labels_to_drop = ['Resource Model Node', 'ODK List Name', 'BI Name', 'Legacy Data Column', 'ODK Value']
drop_condition = df[0].isin(labels_to_drop) | (df[0].isna() & df[1].isna())
df = df.loc[~drop_condition].copy()

# Transform column 8 and 10: replace spaces with underscores and lowercase
df[8] = df[8].str.replace(' ', '_').str.lower()
df[10] = df[10].str.replace(' ', '_').str.lower()

# Replace 'NOT in ODK' with blank in column 11
df[11] = df[11].astype(str).replace('Not in ODK', '')

# Drop unnecessary columns (4 through 7)
df.drop(df.columns[4:7], axis=1, inplace=True)

# Rename columns to meaningful names
df.columns = ["odk_value", "concept_key", "definition", "list_order", "ODK_multi", "list_name", "concept_value", "bulk_import", "ODK_list_name"]

# Save the final DataFrame to CSV, quoting all values
output_csv = os.path.join(data_dir, '1_Processing/excel_thesauri_processed.csv')
df.to_csv(output_csv, index=False, quoting=csv.QUOTE_ALL)

# =======================
# Step 1: Sort thesauri CSV by 'list_name' then 'concept_value'
# =======================
df.sort_values(by=['list_name', 'concept_value'], inplace=True)
df.reset_index(drop=True, inplace=True)

# =======================
# Step 2: Read arches thesauri export
# =======================
arches_path = os.path.join(data_dir, 'arches_thesauri_export.xlsx')
arches_df = pd.read_excel(arches_path)

# =======================
# FORCE INCLUDE IN ARCHES SPREADSHEET - Copy over artefacts_cultural_period* from thesauri_processed.csv
# =======================
thesauri_path = os.path.join(data_dir, '1_Processing/excel_thesauri_processed.csv')
thesauri_df = pd.read_csv(thesauri_path)

# The two list_names we want to copy
forced_list_names = ['artefacts_cultural_period', 'artefacts_cultural_period_certainity']

# Filter thesauri for just these
forced_rows = thesauri_df[thesauri_df['list_name'].isin(forced_list_names)].copy()

# Re-map columns to arches format
forced_rows = pd.DataFrame({
    'list_name': forced_rows['list_name'],
    'parentid': "",  # leave blank
    'concept_value': forced_rows['concept_value'],
    'concept_key': forced_rows['concept_key'],
    'relationshiptype': "narrower",  # constant
    'sortorder': 1,  # constant
    'arches_conceptid': ""  # leave blank
})

# Append these to arches_df
arches_df = pd.concat([arches_df, forced_rows], ignore_index=True)

# =======================
# Step 3: Make a copy of the arches spreadsheet
# =======================
arches_processed_path = os.path.join(data_dir, '1_Processing/arches_thesauri_processed.xlsx')
arches_df.to_excel(arches_processed_path, index=False)

# =======================
# Step 4: Sort the copied arches spreadsheet by 'list_name' then 'concept_value'
# =======================
arches_df.sort_values(by=['list_name', 'concept_value'], inplace=True)
arches_df.reset_index(drop=True, inplace=True)

# =======================
# Step 5: Unique list_name values from thesauri CSV
# =======================
thesauri_unique = pd.DataFrame(df['list_name'].dropna().unique(), columns=['thesauri_list_name'])

# =======================
# Step 6: Unique list_name values from arches processed CSV
# =======================
arches_unique = pd.DataFrame(arches_df['list_name'].dropna().unique(), columns=['arches_list_name'])

# =======================
# Step 7: Exact matches between the two unique lists
# =======================
exact_matches = pd.merge(
    thesauri_unique,
    arches_unique,
    left_on='thesauri_list_name',
    right_on='arches_list_name',
    how='inner'
)
exact_matches['exact_match'] = 'yes'

# =======================
# Step 8: Close matches for thesauri unique values not in exact matches
# =======================

# Identify unmatched values
thesauri_unmatched = thesauri_unique[~thesauri_unique['thesauri_list_name'].isin(exact_matches['thesauri_list_name'])]
arches_unmatched = arches_unique[~arches_unique['arches_list_name'].isin(exact_matches['arches_list_name'])]

# Function to find close match
def find_close(value, choices):
    matches = get_close_matches(value, choices, n=1, cutoff=0.8)  # cutoff=0.8 for similarity
    return matches[0] if matches else pd.NA

# Build DataFrame 4 for thesauri unmatched
list_name_t_nm = thesauri_unmatched.copy()
list_name_t_nm['arches_list_name'] = list_name_t_nm['thesauri_list_name'].apply(lambda x: find_close(x, arches_unmatched['arches_list_name'].tolist()))
list_name_t_nm['close_match'] = list_name_t_nm['arches_list_name'].apply(lambda x: 'yes' if pd.notna(x) else 'no')

# =======================
# Step 9: Close matches for arches unmatched values
# =======================
list_name_a_nm = arches_unmatched.copy()
list_name_a_nm['thesauri_list_name'] = list_name_a_nm['arches_list_name'].apply(lambda x: find_close(x, thesauri_unmatched['thesauri_list_name'].tolist()))
list_name_a_nm['close_match'] = list_name_a_nm['thesauri_list_name'].apply(lambda x: 'yes' if pd.notna(x) else 'no')

# =======================
# Step 10: Create new Excel file with three tabs
# =======================
output_excel_path = os.path.join(data_dir, '2_Comparison/thesauri_arches_list_name_comparison.xlsx')

# Combine thesauri-only and arches-only non-matches into one DataFrame
df_list_name_nm = pd.concat([list_name_t_nm, list_name_a_nm], ignore_index=True)

# Sort so that close matches ("yes") appear first
if "close_match" in df_list_name_nm.columns:
    df_list_name_nm = df_list_name_nm.sort_values(by="close_match", ascending=False)

# Save two sheets instead of three
with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
    exact_matches.to_excel(writer, sheet_name='list_name_matches', index=False)
    df_list_name_nm.to_excel(writer, sheet_name="list_name_nm", index=False)

# Print messages of counts of list names (matchign and not matching), and whether everything matches or not
print("=" * 60)
print('Comparison completed')
print("=" * 60)
print(f"Thesauri unique list_name count: {len(thesauri_unique)}")
print(f"Arches unique list_name count: {len(arches_unique)}")
num_matches = len(exact_matches)
print(f"Number of exact matches between thesauri and arches list_name values: {num_matches}")
print(f"Number of non-matches between thesauri and arches list_name values: {len(df_list_name_nm)}")

if len(df_list_name_nm) > 0:
    print("=" * 60)
    print("⚠️  NOT ALL LISTS MATCH ⚠️")
    print(f"Check the output file for details:\n{output_excel_path}")
    print("=" * 60)
else:
    print("=" * 60)
    print("✅ COMPLETE MATCH! Move onto the next step.")
    print("=" * 60)